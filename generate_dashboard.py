#!/usr/bin/env python3
"""
读取 datset_demo100.xlsx，生成 BioIntelOS 多组学数据门户预览网页 (dashboard.html)。
与 Java 版 DashboardGenerator 输出一致。
"""
import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / "datset_demo100.xlsx"
OUTPUT_PATH = Path(__file__).parent / "dashboard.html"


def read_xlsx(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append({
            "sampleId": str(r[0]) if r[0] else "",
            "cellCount": int(r[1]) if r[1] else 0,
            "species": str(r[2]).strip() if r[2] else "",
            "tissue": str(r[3]).strip() if r[3] else "",
            "disease": str(r[4]).strip() if r[4] else "",
            "omics": str(r[5]).strip() if r[5] else "",
        })
    return rows


def esc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def js(obj):
    return json.dumps(obj, ensure_ascii=False)


def generate_html(data):
    species_set = sorted(set(d["species"] for d in data))
    tissue_set = sorted(set(d["tissue"] for d in data))
    disease_set = sorted(set(d["disease"] for d in data))
    omics_set = sorted(set(d["omics"] for d in data))

    species_count = Counter(d["species"] for d in data)
    omics_count = Counter(d["omics"] for d in data)
    tissue_count = Counter(d["tissue"] for d in data)

    # Sankey links
    st_links = Counter((d["species"], d["tissue"]) for d in data)
    td_links = Counter((d["tissue"], d["disease"]) for d in data)

    # Species x Omics cross-tab
    cross = Counter((d["species"], d["omics"]) for d in data)
    max_cross = max(cross.values()) if cross else 1

    # Tissue bar top 15
    tissue_top = tissue_count.most_common(15)
    tissue_top.reverse()

    # Circle packing data
    cp_children = [{"name": k, "value": v} for k, v in species_count.items()]

    # Omics pie data
    omics_pie = [{"name": k, "value": v} for k, v in omics_count.items()]

    # Heatmap data
    heatmap_data = []
    for si, sp in enumerate(species_set):
        for oi, om in enumerate(omics_set):
            heatmap_data.append([oi, si, cross.get((sp, om), 0)])

    # Sankey
    sankey_nodes_set = set()
    sankey_nodes_set.update(species_set)
    sankey_nodes_set.update(tissue_set)
    sankey_nodes_set.update(disease_set)
    sankey_nodes = [{"name": n} for n in sorted(sankey_nodes_set)]
    sankey_links = []
    for (s, t), v in st_links.items():
        sankey_links.append({"source": s, "target": t, "value": v})
    for (t, d), v in td_links.items():
        sankey_links.append({"source": t, "target": d, "value": v})

    # Table rows
    table_rows = ""
    for d in data:
        table_rows += (
            f'<tr><td>{esc(d["sampleId"])}</td><td>{d["cellCount"]}</td>'
            f'<td>{esc(d["species"])}</td><td>{esc(d["tissue"])}</td>'
            f'<td>{esc(d["disease"])}</td><td>{esc(d["omics"])}</td></tr>\n'
        )

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BioIntelOS - Multi-omics Data Portal</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/d3@7"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f8fafc;color:#334155}}
.container{{max-width:1400px;margin:0 auto;padding:32px}}
.header{{margin-bottom:24px}}
.header h1{{font-size:32px;color:#16a34a;font-weight:700}}
.header p{{color:#94a3b8;font-size:14px;margin-top:4px}}
.stats{{display:grid;grid-template-columns:repeat(5,1fr);gap:16px;margin-bottom:24px}}
.stat-card{{background:#fff;border-radius:16px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08);text-align:center}}
.stat-card .value{{font-size:36px;font-weight:700;color:#16a34a}}
.stat-card .label{{font-size:13px;color:#94a3b8;margin-top:4px}}
.charts{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:20px}}
.chart-card{{background:#fff;border-radius:16px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08)}}
.chart-card h3{{font-size:16px;margin-bottom:4px}}
.chart-card .desc{{font-size:12px;color:#94a3b8;margin-bottom:12px}}
.full-width{{grid-column:1/-1}}
.table-wrap{{background:#fff;border-radius:16px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-top:20px;overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:13px}}
th{{background:#f1f5f9;padding:10px 12px;text-align:left;font-weight:600}}
td{{padding:8px 12px;border-bottom:1px solid #f1f5f9}}
tr:hover td{{background:#f8fafc}}
</style>
</head>
<body>
<div class="container">

<div class="header">
  <h1>BioIntelOS</h1>
  <p>Multi-omics Data Portal</p>
</div>

<div class="stats">
  <div class="stat-card"><div class="value">{len(data)}</div><div class="label">Datasets</div></div>
  <div class="stat-card"><div class="value">{len(species_set)}</div><div class="label">Species</div></div>
  <div class="stat-card"><div class="value">{len(tissue_set)}</div><div class="label">Tissues</div></div>
  <div class="stat-card"><div class="value">{len(disease_set)}</div><div class="label">Diseases</div></div>
  <div class="stat-card"><div class="value">{len(omics_set)}</div><div class="label">Omics</div></div>
</div>

<div class="charts">
  <div class="chart-card">
    <h3>Species Distribution</h3>
    <div class="desc">Circle Packing view</div>
    <svg id="circle-packing" width="100%" height="320"></svg>
  </div>
  <div class="chart-card">
    <h3>Omics Distribution</h3>
    <div class="desc">Data types overview</div>
    <div id="omics-pie" style="height:320px"></div>
  </div>
  <div class="chart-card">
    <h3>Tissue Distribution</h3>
    <div class="desc">Top tissues by dataset count</div>
    <div id="tissue-bar" style="height:320px"></div>
  </div>
  <div class="chart-card">
    <h3>Species × Omics</h3>
    <div class="desc">Cross-tabulation heatmap</div>
    <div id="species-omics" style="height:320px"></div>
  </div>
  <div class="chart-card full-width">
    <h3>Species → Tissue → Disease Sankey</h3>
    <div class="desc">Flow from species through tissue to disease</div>
    <div id="sankey" style="height:500px"></div>
  </div>
</div>

<div class="table-wrap">
  <h3 style="margin-bottom:12px">Dataset Table</h3>
  <table>
    <thead><tr><th>样本ID</th><th>细胞数</th><th>物种</th><th>组织</th><th>疾病</th><th>组学类型</th></tr></thead>
    <tbody>
{table_rows}
    </tbody>
  </table>
</div>

<script>
// Circle Packing (D3)
(function() {{
  var cpData = {{children: {js(cp_children)}}};
  var svg = d3.select('#circle-packing');
  var w = svg.node().getBoundingClientRect().width || 400;
  var h = 320;
  svg.attr('viewBox', '0 0 '+w+' '+h);
  var root = d3.hierarchy(cpData).sum(d => d.value);
  d3.pack().size([w, h]).padding(20)(root);
  var color = d3.scaleOrdinal(['#22c55e','#3b82f6','#f59e0b','#ef4444']);
  var nodes = root.descendants().filter(d => !d.children);
  svg.selectAll('circle').data(nodes).enter().append('circle')
    .attr('cx',d=>d.x).attr('cy',d=>d.y).attr('r',d=>d.r)
    .attr('fill',(d,i)=>color(i)).attr('opacity',0.7);
  svg.selectAll('text').data(nodes).enter().append('text')
    .attr('x',d=>d.x).attr('y',d=>d.y-8).attr('text-anchor','middle')
    .style('font-size','13px').style('font-weight','600').text(d=>d.data.name);
  svg.selectAll('.val').data(nodes).enter().append('text')
    .attr('x',d=>d.x).attr('y',d=>d.y+12).attr('text-anchor','middle')
    .style('font-size','11px').style('fill','#555').text(d=>d.data.value);
}})();

// Omics Pie
echarts.init(document.getElementById('omics-pie')).setOption({{
  tooltip:{{trigger:'item'}},
  color:['#22c55e','#3b82f6','#f59e0b','#ef4444','#8b5cf6'],
  series:[{{type:'pie',radius:['40%','70%'],itemStyle:{{borderRadius:6}},
    label:{{formatter:'{{b}}\\n{{d}}%'}},
    data:{js(omics_pie)}
  }}]
}});

// Tissue Bar
echarts.init(document.getElementById('tissue-bar')).setOption({{
  tooltip:{{}},
  grid:{{left:'25%',right:'10%',top:'5%',bottom:'5%'}},
  xAxis:{{type:'value'}},
  yAxis:{{type:'category',data:{js([t[0] for t in tissue_top])}}},
  series:[{{type:'bar',data:{js([t[1] for t in tissue_top])},
    itemStyle:{{color:'#22c55e',borderRadius:[0,4,4,0]}}}}]
}});

// Species x Omics Heatmap
echarts.init(document.getElementById('species-omics')).setOption({{
  tooltip:{{position:'top'}},
  grid:{{left:'18%',right:'15%',top:'10%',bottom:'18%'}},
  xAxis:{{type:'category',data:{js(omics_set)},axisLabel:{{rotate:30}}}},
  yAxis:{{type:'category',data:{js(species_set)}}},
  visualMap:{{min:0,max:{max_cross},calculable:true,orient:'vertical',right:0,top:'center',
    inRange:{{color:['#f0fdf4','#22c55e','#15803d']}}}},
  series:[{{type:'heatmap',data:{js(heatmap_data)},label:{{show:true}}}}]
}});

// Sankey
echarts.init(document.getElementById('sankey')).setOption({{
  tooltip:{{trigger:'item'}},
  series:[{{type:'sankey',layoutIterations:32,
    emphasis:{{focus:'adjacency'}},
    lineStyle:{{color:'source',curveness:0.5}},
    data:{js(sankey_nodes)},
    links:{js(sankey_links)}
  }}]
}});
</script>

</div>
</body>
</html>"""
    return html


def main():
    data = read_xlsx(XLSX_PATH)
    print(f"读取 {len(data)} 条记录")
    html = generate_html(data)
    OUTPUT_PATH.write_text(html, encoding="utf-8")
    print(f"预览页面已生成: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
