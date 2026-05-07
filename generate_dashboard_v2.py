#!/usr/bin/env python3
"""
BioIntelOS Multi-omics Data Portal — 完整可视化看板生成器 v2
读取 datset_demo100.xlsx，按照详细设计规格生成 dashboard_v2.html。

布局：
  - 顶部 5 张指标卡片（左右分栏：图标+数值+文字 | 趋势折线）
  - 第一行 3 子图：数据规模趋势 | 组学分布环形图 | 物种气泡图
  - 第二行 3 子图：组织条形图 | 疾病条形图 | 数据集信息列表
  - 第三行 1 通栏：桑基图
"""
import json, random, math
from collections import Counter
from pathlib import Path
import openpyxl

XLSX = Path(__file__).parent / "datset_demo100.xlsx"
OUT  = Path(__file__).parent / "dashboard_v2.html"

# ── helpers ──────────────────────────────────────────────
def js(obj):
    return json.dumps(obj, ensure_ascii=False)

def esc(s):
    return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

# ── read data ────────────────────────────────────────────
def read_data():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append(dict(
            sid=str(r[0]) if r[0] else "",
            cells=int(r[1]) if r[1] else 0,
            species=str(r[2]).strip() if r[2] else "",
            tissue=str(r[3]).strip() if r[3] else "",
            disease=str(r[4]).strip() if r[4] else "",
            omics=str(r[5]).strip() if r[5] else "",
        ))
    return rows

# ── sparkline SVG path ───────────────────────────────────
def sparkline_path(seed=42, n=12, w=100, h=36, rising=True):
    rng = random.Random(seed)
    pts = []
    for i in range(n):
        base = (i / (n - 1)) * 0.6 + 0.2 if rising else 0.5
        y = base + rng.uniform(-0.12, 0.12)
        y = max(0.05, min(0.95, y))
        pts.append((round(i * w / (n - 1), 1), round((1 - y) * h, 1)))
    d = "M" + " L".join(f"{x},{y}" for x, y in pts)
    area = d + f" L{w},{h} L0,{h} Z"
    return d, area

# ── SVG icons (simple inline) ────────────────────────────
ICONS = {
    "database": '<svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round"><ellipse cx="12" cy="5" rx="9" ry="3"/><path d="M21 12c0 1.66-4 3-9 3s-9-1.34-9-3"/><path d="M3 5v14c0 1.66 4 3 9 3s9-1.34 9-3V5"/></svg>',
    "testtube": '<svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round"><path d="M9 2v6l-5 10a2 2 0 002 2h12a2 2 0 002-2L15 8V2"/><path d="M9 2h6"/><path d="M5 16h14"/></svg>',
    "cell": '<svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2"><circle cx="12" cy="12" r="9"/><circle cx="12" cy="12" r="3"/><line x1="12" y1="3" x2="12" y2="9"/><line x1="12" y1="15" x2="12" y2="21"/><line x1="3" y1="12" x2="9" y2="12"/><line x1="15" y1="12" x2="21" y2="12"/></svg>',
    "paw": '<svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="#22c55e" stroke-width="2" stroke-linecap="round"><circle cx="11" cy="4" r="2"/><circle cx="4.5" cy="8.5" r="2"/><circle cx="17.5" cy="8.5" r="2"/><circle cx="7" cy="13" r="2"/><circle cx="15" cy="13" r="2"/><path d="M12 17c-2 2-5 3-5 5 0 1 1 2 5 2s5-1 5-2c0-2-3-3-5-5z"/></svg>',
    "omics": '<svg width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="#8b5cf6" stroke-width="2" stroke-linecap="round"><rect x="3" y="3" width="7" height="7" rx="1"/><rect x="14" y="3" width="7" height="7" rx="1"/><rect x="3" y="14" width="7" height="7" rx="1"/><rect x="14" y="14" width="7" height="7" rx="1"/></svg>',
}

# ── tissue icons (unicode) ────────────────────────────────
TISSUE_ICONS = {
    "脑组织":"🧠","肝脏":"🫁","心脏":"❤️","肺组织":"🫁","肾脏":"🫘",
    "肠道":"🔬","结肠":"🔬","皮肤":"🧬","骨髓":"🦴","胰腺":"🧪",
    "乳腺":"🔬","胃":"🧫","脾脏":"🔬","膀胱":"🔬","卵巢":"🔬",
    "甲状腺":"🔬","前列腺":"🔬","子宫":"🔬","食管":"🔬","胚胎":"🧬",
}

# ── HTML head + CSS ───────────────────────────────────────
HTML_HEAD = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>BioIntelOS — Multi-omics Data Portal</title>
<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/d3@7"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{
  font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Hiragino Sans GB',
    'Microsoft YaHei',Roboto,Helvetica,Arial,sans-serif;
  background:#fff;color:#334155;line-height:1.5;
}
.container{max-width:1440px;margin:0 auto;padding:36px 40px}

/* ── brand header ── */
.brand{margin-bottom:28px}
.brand h1{font-size:26px;font-weight:700;color:#16a34a;letter-spacing:-0.5px}
.brand h1 span{color:#334155}
.brand p{font-size:13px;color:#94a3b8;margin-top:2px}

/* ── stat cards row ── */
.cards-row{
  display:grid;grid-template-columns:repeat(5,1fr);gap:16px;margin-bottom:28px;
}
.stat-card{
  background:#fff;border:1px solid #e2e8f0;border-radius:14px;
  padding:18px 20px;display:flex;align-items:center;gap:12px;
  transition:box-shadow .2s;
}
.stat-card:hover{box-shadow:0 4px 16px rgba(0,0,0,0.06)}
.sc-left{flex:0 0 60%;display:flex;flex-direction:column;gap:2px}
.sc-icon{margin-bottom:4px}
.sc-value{font-size:28px;font-weight:700;line-height:1.1}
.sc-title{font-size:13px;font-weight:600;color:#475569}
.sc-sub{font-size:11px;color:#94a3b8}
.sc-right{flex:1;display:flex;align-items:flex-end}
.sc-right svg{width:100%;height:36px}

/* ── chart grid ── */
.chart-grid{
  display:grid;grid-template-columns:repeat(3,1fr);gap:20px;
}
.chart-card{
  background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:20px;
  transition:box-shadow .2s;
}
.chart-card:hover{box-shadow:0 4px 16px rgba(0,0,0,0.06)}
.chart-card h3{font-size:15px;font-weight:700;color:#1e293b;margin-bottom:12px}
.full-width{grid-column:1/-1}

/* ── list card ── */
.list-card{overflow:hidden}
.list-cols{display:grid;grid-template-columns:1fr 1fr;gap:20px}
.list-col-title{font-size:13px;font-weight:700;color:#475569;margin-bottom:10px;
  padding-bottom:6px;border-bottom:1px solid #f1f5f9}
.list-item{display:flex;align-items:center;gap:6px;padding:5px 0;font-size:12px;flex-wrap:wrap}
.li-dot{width:6px;height:6px;border-radius:50%;background:#22c55e;flex-shrink:0}
.li-rank{font-weight:700;color:#22c55e;min-width:24px}
.li-name{color:#475569;flex:1;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.li-tag{background:#f0fdf4;color:#16a34a;font-size:10px;padding:1px 7px;border-radius:8px;
  font-weight:600;white-space:nowrap}
.li-cite{font-size:11px;color:#94a3b8;white-space:nowrap}

/* ── tissue icon grid ── */
.tissue-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:16px 12px;padding:8px 0}
.tissue-cell{display:flex;flex-direction:column;align-items:center;gap:6px;text-align:center}
.tissue-icon-wrap{width:48px;height:48px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:22px}
.tissue-name{font-size:12px;color:#475569;font-weight:500;line-height:1.2}
.tissue-pct{font-size:11px;color:#64748b}

/* ── circle packing ── */
#circle-packing{display:block;margin:0 auto}
</style>
</head>
<body>
<div class="container">
<div class="brand">
  <h1>BioIntel<span>OS</span></h1>
  <p>Multi-omics Data Portal · 多组学数据门户</p>
</div>
'''

# ── main generate ─────────────────────────────────────────
def generate(data):
    n = len(data)
    total_cells = sum(d["cells"] for d in data)
    species_set = sorted(set(d["species"] for d in data))
    tissue_set  = sorted(set(d["tissue"]  for d in data))
    disease_set = sorted(set(d["disease"] for d in data))
    omics_set   = sorted(set(d["omics"]   for d in data))

    species_cnt = Counter(d["species"] for d in data)
    omics_cnt   = Counter(d["omics"]   for d in data)
    tissue_cnt  = Counter(d["tissue"]  for d in data)
    disease_cnt = Counter(d["disease"] for d in data)

    # sankey
    st = Counter((d["species"], d["tissue"])  for d in data)
    td = Counter((d["tissue"],  d["disease"]) for d in data)
    sankey_nodes = sorted(set(species_set) | set(tissue_set) | set(disease_set))
    sankey_links = []
    for (s,t),v in st.items(): sankey_links.append({"source":s,"target":t,"value":v})
    for (t,d),v in td.items(): sankey_links.append({"source":t,"target":d,"value":v})

    # synthetic trend data (6 years)
    years = ["2019","2020","2021","2022","2023","2024"]
    ds_trend  = [12, 25, 38, 55, 78, 100]
    smp_trend = [15, 30, 42, 60, 82, 100]

    # top 10 tissues / diseases
    tissue_top10  = tissue_cnt.most_common(10)
    disease_top10 = disease_cnt.most_common(10)

    # bubble data for species
    bubble_data = []
    for i, (sp, cnt) in enumerate(species_cnt.most_common()):
        avg_cells = sum(d["cells"] for d in data if d["species"]==sp) / cnt
        bubble_data.append([i * 25 + 15, avg_cells / 100, cnt, sp])

    # stat cards config
    cards = [
        ("database", "#22c55e", str(n),           "数据集总数", "持续增长中",   1),
        ("testtube", "#22c55e", str(n),            "样本总数",   "来自多个研究", 2),
        ("cell",     "#22c55e", f"{total_cells:,}","细胞总数",   "单细胞数据",   3),
        ("paw",      "#22c55e", str(len(species_set)),"物种数量","覆盖广泛",    4),
        ("omics",    "#8b5cf6", str(len(omics_set)),"数据类型（组学）","多组学整合",5),
    ]

    # ── build HTML ────────────────────────────────────────
    H = []
    H.append(HTML_HEAD)

    # --- stat cards ---
    H.append('<div class="cards-row">')
    for icon_key, color, value, title, sub, seed in cards:
        line_d, area_d = sparkline_path(seed=seed, rising=True)
        fill_c = color.replace(")", ",0.12)").replace("#","") if "#" in color else color
        # convert hex to rgba for fill
        r_,g_,b_ = int(color[1:3],16), int(color[3:5],16), int(color[5:7],16)
        fill_rgba = f"rgba({r_},{g_},{b_},0.10)"
        H.append(f'''<div class="stat-card">
  <div class="sc-left">
    <div class="sc-icon">{ICONS[icon_key]}</div>
    <div class="sc-value" style="color:{color}">{value}</div>
    <div class="sc-title">{title}</div>
    <div class="sc-sub">{sub}</div>
  </div>
  <div class="sc-right">
    <svg viewBox="0 0 100 36" preserveAspectRatio="none">
      <path d="{area_d}" fill="{fill_rgba}" />
      <path d="{line_d}" fill="none" stroke="{color}" stroke-width="2"/>
    </svg>
  </div>
</div>''')
    H.append('</div>')

    # --- chart grid ---
    H.append('<div class="chart-grid">')

    # Row 1, Col 1: 数据规模趋势
    H.append('''<div class="chart-card">
  <h3>数据规模（数据量）</h3>
  <div id="trend" style="height:300px"></div>
</div>''')

    # Row 1, Col 2: 组学分布环形图
    H.append('''<div class="chart-card">
  <h3>数据类型（组学）分布</h3>
  <div id="omics-donut" style="height:300px"></div>
</div>''')

    # Row 1, Col 3: 物种 Circle Packing (D3)
    H.append('''<div class="chart-card">
  <h3>物种分布（Top 10）</h3>
  <svg id="circle-packing" width="380" height="300"></svg>
</div>''')

    # Row 2, Col 1: 组织分布 (2×5 icon grid)
    tissue_top10_grid = tissue_cnt.most_common(9)  # top 9 + "其他"
    other_cnt = sum(v for k,v in tissue_cnt.items() if k not in dict(tissue_top10_grid))
    total_tissue = sum(tissue_cnt.values())

    # icon + color mapping for tissues
    tissue_icon_map = [
        ("脑组织","🧠","#dcfce7"), ("肝脏","🫁","#dbeafe"), ("心脏","❤️","#fce7f3"),
        ("肺组织","🫁","#f0fdf4"), ("肾脏","🫘","#fef3c7"), ("肠道","🔬","#ede9fe"),
        ("结肠","🔬","#e0f2fe"), ("皮肤","🧬","#fef9c3"), ("骨髓","🦴","#fee2e2"),
        ("胰腺","🧪","#f0fdf4"), ("乳腺","🔬","#fce7f3"), ("胃","🧫","#dbeafe"),
        ("脾脏","🔬","#ede9fe"), ("膀胱","🔬","#e0f2fe"), ("卵巢","🔬","#fce7f3"),
        ("甲状腺","🔬","#fef3c7"), ("前列腺","🔬","#dbeafe"), ("子宫","🔬","#fce7f3"),
        ("食管","🔬","#fee2e2"), ("胚胎","🧬","#f0fdf4"),
    ]
    tissue_icon_dict = {t[0]: (t[1], t[2]) for t in tissue_icon_map}

    H.append('<div class="chart-card">')
    H.append('  <h3>组织分布（Top 10）</h3>')
    H.append('  <p style="font-size:12px;color:#64748b;margin-bottom:14px">数据覆盖多种组织类别，以部分类别最为丰富</p>')
    H.append('  <div class="tissue-grid">')
    for tname, tcnt in tissue_top10_grid:
        pct = round(tcnt / total_tissue * 100, 1)
        icon, bg = tissue_icon_dict.get(tname, ("🔬", "#f0fdf4"))
        H.append(f'''    <div class="tissue-cell">
      <div class="tissue-icon-wrap" style="background:{bg}">{icon}</div>
      <div class="tissue-name">{esc(tname)}</div>
      <div class="tissue-pct">{pct}%</div>
    </div>''')
    # "其他" cell
    other_pct = round(other_cnt / total_tissue * 100, 1)
    H.append(f'''    <div class="tissue-cell">
      <div class="tissue-icon-wrap" style="background:#f1f5f9">
        <svg width="20" height="20" viewBox="0 0 20 20"><circle cx="4" cy="10" r="2" fill="#94a3b8"/><circle cx="10" cy="10" r="2" fill="#94a3b8"/><circle cx="16" cy="10" r="2" fill="#94a3b8"/></svg>
      </div>
      <div class="tissue-name">其他</div>
      <div class="tissue-pct">{other_pct}%</div>
    </div>''')
    H.append('  </div>')
    H.append('</div>')

    # Row 2, Col 2: 疾病分布
    H.append('''<div class="chart-card">
  <h3>疾病类型分布（Top 10）</h3>
  <div id="disease-bar" style="height:300px"></div>
</div>''')

    # Row 2, Col 3: 数据集信息列表
    H.append('<div class="chart-card list-card">')
    H.append('<h3>最新更新 &amp; 最常引用数据集</h3>')
    H.append('<div class="list-cols">')
    # left: latest
    H.append('<div class="list-col"><div class="list-col-title">最新更新</div>')
    for d in reversed(data[-6:]):
        H.append(f'''<div class="list-item">
  <span class="li-dot"></span>
  <span class="li-name">{esc(d["sid"])} · {esc(d["tissue"])} · {esc(d["disease"])}</span>
  <span class="li-tag">{esc(d["omics"])}</span>
</div>''')
    H.append('</div>')
    # right: most cited (synthetic)
    H.append('<div class="list-col"><div class="list-col-title">最常引用</div>')
    rng = random.Random(99)
    cited = sorted(data, key=lambda x: x["cells"], reverse=True)[:6]
    for rank, d in enumerate(cited, 1):
        cite_n = rng.randint(80, 320)
        H.append(f'''<div class="list-item">
  <span class="li-rank">#{rank}</span>
  <span class="li-name">{esc(d["sid"])} · {esc(d["tissue"])} · {esc(d["disease"])}</span>
  <span class="li-cite">{cite_n} 引用</span>
  <span class="li-tag">{esc(d["omics"])}</span>
</div>''')
    H.append('</div></div></div>')

    # Row 3: Sankey (full width)
    H.append('''<div class="chart-card full-width">
  <h3>物种 - 组织 - 疾病分布关系</h3>
  <div id="sankey" style="height:520px"></div>
</div>''')

    H.append('</div>') # end chart-grid
    H.append('</div>') # end container

    # ── JavaScript ────────────────────────────────────────
    H.append('<script>')

    # 1. Trend chart
    H.append(f'''
var trend=echarts.init(document.getElementById('trend'));
trend.setOption({{
  tooltip:{{trigger:'axis'}},
  legend:{{data:['数据集数量','样本数量'],top:0,right:0,textStyle:{{fontSize:11}}}},
  grid:{{left:'10%',right:'10%',top:'15%',bottom:'12%'}},
  xAxis:{{type:'category',data:{js(years)},boundaryGap:false}},
  yAxis:[
    {{type:'value',name:'数据集',splitLine:{{lineStyle:{{type:'dashed',color:'#f0f0f0'}}}}}},
    {{type:'value',name:'样本数',splitLine:{{show:false}}}}
  ],
  series:[
    {{name:'数据集数量',type:'line',smooth:true,symbol:'circle',symbolSize:6,
      lineStyle:{{color:'#22c55e',width:2.5}},
      areaStyle:{{color:new echarts.graphic.LinearGradient(0,0,0,1,[
        {{offset:0,color:'rgba(34,197,94,0.25)'}},{{offset:1,color:'rgba(34,197,94,0.02)'}}
      ])}},
      itemStyle:{{color:'#22c55e'}},
      data:{js(ds_trend)}}},
    {{name:'样本数量',type:'line',smooth:true,yAxisIndex:1,symbol:'circle',symbolSize:6,
      lineStyle:{{color:'#3b82f6',width:2.5}},
      areaStyle:{{color:new echarts.graphic.LinearGradient(0,0,0,1,[
        {{offset:0,color:'rgba(59,130,246,0.20)'}},{{offset:1,color:'rgba(59,130,246,0.02)'}}
      ])}},
      itemStyle:{{color:'#3b82f6'}},
      data:{js(smp_trend)}}}
  ]
}});''')

    # 2. Omics donut
    omics_pie = [{"name":k,"value":v} for k,v in omics_cnt.most_common()]
    H.append(f'''
var donut=echarts.init(document.getElementById('omics-donut'));
donut.setOption({{
  tooltip:{{trigger:'item',formatter:'{{b}}: {{c}} ({{d}}%)'}},
  legend:{{orient:'vertical',right:10,top:'middle',textStyle:{{fontSize:12}}}},
  color:['#22c55e','#3b82f6','#8b5cf6','#f59e0b','#ef4444'],
  series:[{{
    type:'pie',radius:['42%','72%'],center:['40%','50%'],
    itemStyle:{{borderRadius:6,borderColor:'#fff',borderWidth:2}},
    label:{{show:false}},
    emphasis:{{label:{{show:true,fontSize:14,fontWeight:'bold'}}}},
    data:{js(omics_pie)}
  }}]
}});''')

    # 3. Species Circle Packing (D3)
    cp_children = [{"name":k,"value":v} for k,v in species_cnt.most_common()]
    H.append(f'''
// Circle Packing (D3)
(function(){{
  var cpData = {{children:{js(cp_children)}}};
  var svg = d3.select('#circle-packing');
  var w = 380, h = 300;
  svg.attr('viewBox','0 0 '+w+' '+h);
  var root = d3.hierarchy(cpData).sum(d=>d.value);
  d3.pack().size([w,h]).padding(8)(root);
  var color = d3.scaleOrdinal(['#86efac','#93c5fd','#c4b5fd','#fde68a']);
  var nodes = root.descendants().filter(d=>!d.children);
  svg.selectAll('circle').data(nodes).enter().append('circle')
    .attr('cx',d=>d.x).attr('cy',d=>d.y).attr('r',d=>d.r)
    .attr('fill',(d,i)=>color(i)).attr('opacity',0.8)
    .attr('stroke',(d,i)=>d3.color(color(i)).darker(0.3)).attr('stroke-width',1.5);
  svg.selectAll('.lbl').data(nodes).enter().append('text')
    .attr('x',d=>d.x).attr('y',d=>d.y-6).attr('text-anchor','middle')
    .style('font-size','13px').style('font-weight','700').style('fill','#1e293b')
    .text(d=>d.data.name);
  svg.selectAll('.val').data(nodes).enter().append('text')
    .attr('x',d=>d.x).attr('y',d=>d.y+14).attr('text-anchor','middle')
    .style('font-size','11px').style('fill','#475569')
    .text(d=>d.data.value);
}})();
''')

    # 4. (Tissue grid is pure HTML, no JS needed)

    # 5. Disease bar
    d10_names = [d[0] for d in disease_top10]
    d10_vals  = [d[1] for d in disease_top10]
    d10_names.reverse(); d10_vals.reverse()
    H.append(f'''
var dbar=echarts.init(document.getElementById('disease-bar'));
dbar.setOption({{
  tooltip:{{}},
  grid:{{left:'30%',right:'12%',top:'5%',bottom:'5%'}},
  xAxis:{{type:'value',splitLine:{{lineStyle:{{type:'dashed',color:'#f0f0f0'}}}}}},
  yAxis:{{type:'category',data:{js(d10_names)},axisLabel:{{fontSize:12}}}},
  series:[{{type:'bar',data:{js(d10_vals)},barWidth:16,
    itemStyle:{{borderRadius:[0,4,4,0],
      color:new echarts.graphic.LinearGradient(0,0,1,0,[
        {{offset:0,color:'#16a34a'}},{{offset:1,color:'#4ade80'}}
      ])
    }},
    label:{{show:true,position:'right',fontSize:11,color:'#64748b'}}
  }}]
}});''')

    # 6. Sankey
    H.append(f'''
var sk=echarts.init(document.getElementById('sankey'));
sk.setOption({{
  tooltip:{{trigger:'item'}},
  series:[{{
    type:'sankey',layoutIterations:32,nodeWidth:20,nodeGap:10,
    emphasis:{{focus:'adjacency'}},
    lineStyle:{{color:'source',curveness:0.5,opacity:0.35}},
    label:{{fontSize:11}},
    data:{js([{"name":n} for n in sankey_nodes])},
    links:{js(sankey_links)}
  }}]
}});''')

    # Resize handler
    H.append('''
window.addEventListener('resize',function(){
  [trend,donut,dbar,sk].forEach(function(c){c.resize()});
});''')

    H.append('</script>\n</body>\n</html>')
    return "\n".join(H)

def main():
    data = read_data()
    print(f"读取 {len(data)} 条记录")
    html = generate(data)
    OUT.write_text(html, encoding="utf-8")
    print(f"看板已生成: {OUT}")

if __name__ == "__main__":
    main()
